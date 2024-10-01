import json
import time
import sys

from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver

# 셀레눔 드라이버와 스크롤 횟수를 받아서 스크롤을 내리는 함수 
def scroll_page(driver, scroll_count):  # 스크롤 횟수만큼 스크롤을 내림
    for _ in range(scroll_count):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(2)


# BeautifulSoup로 리뷰를 추출하는 함수
def extract_reviews(html):
    soup=BeautifulSoup(html,'html.parser')

    # 리뷰가 들어 있는 html 태그를 찾아서 저장 
    reviews = soup.find_all('div', class_='css-166s55a')
    review_list=[]

    for review in reviews:
        review_texts=review.find_all('p',class_='content-text css-c92dc4')
        review_text=""
        for text in review_texts:
            review_text += text.get_text(strip=True)+" "

        # 별점 추출 (리뷰 하나하나에서 별점 요소 찾기)
        stars = review.find_all('svg', class_='css-1mhwecd')  # 꽉 찬 별의 클래스 사용
        rating_value = 0
        for star in stars:
            path = star.find('path')
            if path and path.get('fill') == '#FDBD00': # 별이 꽉 찼을 때
                rating_value += 1

        # 리뷰와 별점을 리스트에 추가
        review_list.append({'review': review_text.strip(), 'rating': f"{rating_value}점"})
    
    return review_list


# 리뷰를 엑셀로 저장하는 함수
def save_reviews(name, review_list): 
    file = openpyxl.Workbook() 
    sheet = file.active 
    sheet.title = name 

    sheet.cell(row=1,column=1,value="Review")
    sheet.cell(row=1,column=2,value="Rating")

    for i, data in enumerate(review_list, start=2):
        sheet.cell(row=i,column=1,value=data['review'])
        sheet.cell(row=i,column=2,value=data['rating'])

    file.save(f"{name}_reviews.xlsx")
    return
    

def crawl_yanolja_reviews(name, url):
    # 드라이버 
    driver = webdriver.Chrome()
    driver.get(url)
    time.sleep(3)

    # 스크롤 횟수
    scroll_count = 20
    # 스크롤 내리기 함수 실행
    scroll_page(driver, scroll_count)

    # 리뷰 추출
    html = driver.page_source
    # 리뷰 추출 함수 실행
    review_list = extract_reviews(html)

    # 리뷰 저장 함수 실행 
    save_reviews(name, review_list)
    driver.quit()

if __name__ == '__main__':
    name="해운대 씨클라우드 호텔 앤 레지던스"
    url="https://www.yanolja.com/reviews/domestic/1000095499"
    crawl_yanolja_reviews(name=name, url=url)
